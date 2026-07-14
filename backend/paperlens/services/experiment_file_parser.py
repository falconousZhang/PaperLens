from __future__ import annotations

import csv
import datetime as dt
import io
import math
import re
import unicodedata
import zipfile
from dataclasses import dataclass, field
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree

from paperlens.core.enums import ExperimentFileType


_COLUMNS_INFO_VERSION = 1
_MAX_FILE_SIZE = 20 * 1024 * 1024
_MAX_ROWS = 100000
_MAX_COLUMNS = 256
_MAX_ZIP_ENTRIES = 5000
_MAX_UNCOMPRESSED_SIZE = 200 * 1024 * 1024
_MAX_COMPRESSION_RATIO = 100.0
_MAX_COL_NAME_LENGTH = 128
_MAX_FILENAME_LENGTH = 255
_CSV_DELIMITERS = (",", ";", "\t")
_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class ParseError(Exception):
    def __init__(self, message: str, kind: str = "content"):
        super().__init__(message)
        self.kind = kind


@dataclass(frozen=True)
class ParseResult:
    file_type: ExperimentFileType
    row_count: int
    column_count: int
    columns_info: dict
    encoding: str | None
    delimiter: str | None
    sheet_name: str | None


@dataclass
class _ColumnAccumulator:
    kinds: set[str] = field(default_factory=set)
    null_count: int = 0

    def observe(self, value) -> None:
        if value is None:
            self.null_count += 1
            return
        if isinstance(value, bool):
            self.kinds.add("boolean")
        elif isinstance(value, int):
            self.kinds.add("integer")
        elif isinstance(value, float):
            if math.isfinite(value):
                self.kinds.add("integer" if value.is_integer() else "float")
            else:
                self.kinds.add("string")
        elif isinstance(value, (dt.datetime, dt.date)):
            self.kinds.add("datetime")
        else:
            self.kinds.add("string")

    def dtype(self) -> str:
        if not self.kinds:
            return "empty"
        if self.kinds == {"integer"}:
            return "integer"
        if self.kinds <= {"integer", "float"}:
            return "float"
        if self.kinds == {"boolean"}:
            return "boolean"
        if self.kinds == {"datetime"}:
            return "datetime"
        return "string"


def _is_control_char(value: str) -> bool:
    return any(unicodedata.category(char) == "Cc" for char in value)


def validate_filename_and_type(filename: str) -> tuple[str, ExperimentFileType]:
    basename = filename.replace("\\", "/").rsplit("/", 1)[-1]
    if not basename or len(basename) > _MAX_FILENAME_LENGTH or _is_control_char(basename):
        raise ParseError("invalid filename", "type")
    extension = Path(basename).suffix.casefold()
    mapping = {
        ".csv": ExperimentFileType.CSV,
        ".xlsx": ExperimentFileType.XLSX,
        ".xls": ExperimentFileType.XLS,
    }
    if extension not in mapping:
        raise ParseError("unsupported file extension", "type")
    return basename, mapping[extension]


def _validate_source_path(source_path: str | Path) -> Path:
    path = Path(source_path)
    try:
        size = path.stat().st_size
    except OSError as exc:
        raise ParseError("source file is unavailable") from exc
    if not path.is_file():
        raise ParseError("source path is not a file")
    if size < 1:
        raise ParseError("source file is empty", "type")
    if size > _MAX_FILE_SIZE:
        raise ParseError("source file exceeds size limit", "size")
    return path


def validate_container(source_path: str | Path, file_type: ExperimentFileType) -> Path:
    path = _validate_source_path(source_path)
    try:
        with path.open("rb") as stream:
            magic = stream.read(8)
    except OSError as exc:
        raise ParseError("source file cannot be read") from exc
    if file_type == ExperimentFileType.CSV:
        if magic.startswith((b"PK\x03\x04", b"PK\x05\x06", b"\x1f\x8b")) or magic == _OLE_MAGIC:
            raise ParseError("CSV container does not match extension", "type")
    elif file_type == ExperimentFileType.XLSX:
        if not magic.startswith(b"PK\x03\x04"):
            raise ParseError("XLSX container does not match extension", "type")
    elif file_type == ExperimentFileType.XLS:
        if magic != _OLE_MAGIC:
            raise ParseError("XLS container does not match extension", "type")
    else:
        raise ParseError("unsupported file type", "type")
    return path


def _validate_column_names(values: list) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for value in values:
        name = ("" if value is None else str(value)).lstrip("\ufeff").strip()
        if not name or len(name) > _MAX_COL_NAME_LENGTH or _is_control_char(name):
            raise ParseError("invalid column name")
        if name in seen:
            raise ParseError("duplicate column name")
        seen.add(name)
        names.append(name)
    return names


def _build_columns_info(
    names: list[str],
    accumulators: list[_ColumnAccumulator],
    encoding: str | None,
    delimiter: str | None,
    sheet_name: str | None,
) -> dict:
    return {
        "version": _COLUMNS_INFO_VERSION,
        "encoding": encoding,
        "delimiter": delimiter,
        "sheet_name": sheet_name,
        "columns": [
            {
                "name": name,
                "dtype": accumulator.dtype(),
                "nullable": accumulator.null_count > 0,
                "null_count": accumulator.null_count,
            }
            for name, accumulator in zip(names, accumulators, strict=True)
        ],
    }


def _decode_csv(raw: bytes) -> tuple[str, str]:
    if raw.startswith(b"\xef\xbb\xbf"):
        try:
            return raw[3:].decode("utf-8"), "utf-8-sig"
        except UnicodeDecodeError as exc:
            raise ParseError("invalid UTF-8 BOM CSV") from exc
    try:
        return raw.decode("utf-8"), "utf-8"
    except UnicodeDecodeError:
        try:
            return raw.decode("gb18030"), "gb18030"
        except UnicodeDecodeError as exc:
            raise ParseError("CSV encoding is unsupported") from exc


def _delimiter_profile(text: str, delimiter: str) -> tuple[bool, int, int]:
    width = 0
    row_count = 0
    try:
        reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter, strict=True)
        for row in reader:
            row_count += 1
            if not row:
                return False, 0, row_count
            if width == 0:
                width = len(row)
            elif len(row) != width:
                return False, width, row_count
    except csv.Error:
        return False, 0, row_count
    return row_count > 0, width, row_count


def _detect_delimiter(text: str) -> str:
    profiles = {delimiter: _delimiter_profile(text, delimiter) for delimiter in _CSV_DELIMITERS}
    candidates = [
        delimiter
        for delimiter, (stable, width, _) in profiles.items()
        if stable and width > 1
    ]
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        raise ParseError("CSV delimiter is ambiguous")
    if all(stable and width == 1 for stable, width, _ in profiles.values()):
        return ","
    raise ParseError("CSV delimiter or row width is inconsistent")


def _infer_csv_value(value: str):
    if value == "":
        return None
    lowered = value.casefold()
    if lowered in {"true", "false"}:
        return lowered == "true"
    if re.fullmatch(r"[+-]?\d+", value):
        digits = value.lstrip("+-")
        if len(digits) == 1 or not digits.startswith("0"):
            try:
                return int(value)
            except ValueError:
                pass
    try:
        number = float(value)
    except ValueError:
        number = None
    if number is not None and math.isfinite(number):
        return number
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[T ]\d{2}:\d{2}(?::\d{2}(?:\.\d{1,6})?)?(?:Z|[+-]\d{2}:\d{2})?)?", value):
        try:
            normalized = value[:-1] + "+00:00" if value.endswith("Z") else value
            return dt.datetime.fromisoformat(normalized) if "T" in normalized or " " in normalized else dt.date.fromisoformat(normalized)
        except ValueError:
            pass
    return value


def parse_csv(source_path: str | Path) -> ParseResult:
    path = validate_container(source_path, ExperimentFileType.CSV)
    try:
        raw = path.read_bytes()
    except OSError as exc:
        raise ParseError("CSV cannot be read") from exc
    text, encoding = _decode_csv(raw)
    if "\x00" in text:
        raise ParseError("CSV contains NUL")
    delimiter = _detect_delimiter(text)
    try:
        reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter, strict=True)
        header = next(reader, None)
        if header is None:
            raise ParseError("CSV has no header")
        column_count = len(header)
        if not 1 <= column_count <= _MAX_COLUMNS:
            raise ParseError("CSV column count exceeds limit", "size")
        names = _validate_column_names(header)
        accumulators = [_ColumnAccumulator() for _ in names]
        row_count = 0
        for row in reader:
            row_count += 1
            if row_count > _MAX_ROWS:
                raise ParseError("CSV row count exceeds limit", "size")
            if len(row) != column_count:
                raise ParseError("CSV row width is inconsistent")
            for accumulator, value in zip(accumulators, row, strict=True):
                accumulator.observe(_infer_csv_value(value))
    except csv.Error as exc:
        raise ParseError("CSV syntax is invalid") from exc
    if row_count == 0:
        raise ParseError("CSV contains no data rows")
    columns_info = _build_columns_info(names, accumulators, encoding, delimiter, None)
    return ParseResult(
        file_type=ExperimentFileType.CSV,
        row_count=row_count,
        column_count=column_count,
        columns_info=columns_info,
        encoding=encoding,
        delimiter=delimiter,
        sheet_name=None,
    )


def _zip_ratio_exceeded(compressed: int, uncompressed: int) -> bool:
    if uncompressed == 0:
        return False
    if compressed == 0:
        return True
    return uncompressed / compressed > _MAX_COMPRESSION_RATIO


def _normalized_zip_name(name: str) -> str:
    normalized = name.replace("\\", "/")
    parts = PurePosixPath(normalized).parts
    if (
        "\\" in name
        or normalized.startswith("/")
        or re.match(r"^[A-Za-z]:", normalized)
        or ".." in parts
    ):
        raise ParseError("XLSX contains an unsafe entry path")
    return normalized


def _xml_has_element(stream, local_name: str) -> bool:
    try:
        for _, element in ElementTree.iterparse(stream, events=("start",)):
            if element.tag.rsplit("}", 1)[-1].casefold() == local_name.casefold():
                return True
    except ElementTree.ParseError as exc:
        raise ParseError("XLSX contains malformed XML") from exc
    return False


def _xml_has_external_relationship(stream) -> bool:
    try:
        for _, element in ElementTree.iterparse(stream, events=("start",)):
            if element.tag.rsplit("}", 1)[-1] == "Relationship":
                if element.attrib.get("TargetMode", "").casefold() == "external":
                    return True
    except ElementTree.ParseError as exc:
        raise ParseError("XLSX contains malformed relationship XML") from exc
    return False


def _check_xlsx_zip_safety(zf: zipfile.ZipFile) -> None:
    entries = zf.infolist()
    if len(entries) > _MAX_ZIP_ENTRIES:
        raise ParseError("XLSX entry count exceeds limit", "size")
    total_compressed = 0
    total_uncompressed = 0
    seen: set[str] = set()
    normalized_names: dict[str, zipfile.ZipInfo] = {}
    for entry in entries:
        normalized = _normalized_zip_name(entry.filename)
        folded = normalized.casefold()
        if folded in seen:
            raise ParseError("XLSX contains duplicate entries")
        seen.add(folded)
        normalized_names[normalized] = entry
        if entry.flag_bits & 0x1:
            raise ParseError("XLSX contains an encrypted entry")
        total_compressed += entry.compress_size
        total_uncompressed += entry.file_size
        if _zip_ratio_exceeded(entry.compress_size, entry.file_size):
            raise ParseError("XLSX entry compression ratio exceeds limit", "size")
        if (
            folded == "xl/vbaproject.bin"
            or folded.startswith("xl/macrosheets/")
            or folded.startswith("xl/dialogsheets/")
        ):
            raise ParseError("XLSX contains macros")
        if folded.startswith("xl/externallinks/"):
            raise ParseError("XLSX contains external links")
        if folded.startswith(("xl/embeddings/", "xl/oleobjects/", "xl/activex/")):
            raise ParseError("XLSX contains embedded objects")
    if total_uncompressed > _MAX_UNCOMPRESSED_SIZE:
        raise ParseError("XLSX uncompressed size exceeds limit", "size")
    if _zip_ratio_exceeded(total_compressed, total_uncompressed):
        raise ParseError("XLSX total compression ratio exceeds limit", "size")
    required = {"[Content_Types].xml", "xl/workbook.xml"}
    if not required.issubset(normalized_names):
        raise ParseError("XLSX required parts are missing", "type")
    try:
        content_types = ElementTree.fromstring(zf.read(normalized_names["[Content_Types].xml"]))
    except (ElementTree.ParseError, KeyError) as exc:
        raise ParseError("XLSX content types are invalid", "type") from exc
    for element in content_types.iter():
        content_type = element.attrib.get("ContentType", "").casefold()
        if "macroenabled" in content_type or "vbaproject" in content_type:
            raise ParseError("XLSX content types declare macros")
    for normalized, entry in normalized_names.items():
        folded = normalized.casefold()
        if folded.startswith("xl/worksheets/") and folded.endswith(".xml"):
            with zf.open(entry) as stream:
                if _xml_has_element(stream, "f"):
                    raise ParseError("XLSX contains formula cells")
        if folded.startswith("xl/") and folded.endswith(".rels"):
            with zf.open(entry) as stream:
                if _xml_has_external_relationship(stream):
                    raise ParseError("XLSX contains an external relationship")


def _worksheet_has_values(worksheet) -> bool:
    for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if index > _MAX_ROWS + 1:
            raise ParseError("XLSX row count exceeds limit", "size")
        if any(value is not None for value in row):
            return True
    return False


def parse_xlsx(source_path: str | Path) -> ParseResult:
    path = validate_container(source_path, ExperimentFileType.XLSX)
    try:
        with zipfile.ZipFile(path) as archive:
            _check_xlsx_zip_safety(archive)
    except zipfile.BadZipFile as exc:
        raise ParseError("XLSX is not a valid ZIP container", "type") from exc
    try:
        import openpyxl
    except ImportError as exc:
        raise ParseError("XLSX parser is unavailable") from exc
    try:
        workbook = openpyxl.load_workbook(
            filename=str(path),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise ParseError("XLSX workbook cannot be parsed") from exc
    try:
        non_empty = [name for name in workbook.sheetnames if _worksheet_has_values(workbook[name])]
        if len(non_empty) != 1:
            raise ParseError("XLSX must contain exactly one non-empty worksheet")
        sheet_name = non_empty[0]
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = list(next(rows, ()))
        while header and header[-1] is None:
            header.pop()
        column_count = len(header)
        if not 1 <= column_count <= _MAX_COLUMNS:
            raise ParseError("XLSX column count exceeds limit", "size")
        names = _validate_column_names(header)
        accumulators = [_ColumnAccumulator() for _ in names]
        row_count = 0
        for row in rows:
            if any(value is not None for value in row[column_count:]):
                raise ParseError("XLSX row width exceeds header width")
            row_count += 1
            if row_count > _MAX_ROWS:
                raise ParseError("XLSX row count exceeds limit", "size")
            values = list(row[:column_count])
            values.extend([None] * (column_count - len(values)))
            for accumulator, value in zip(accumulators, values, strict=True):
                accumulator.observe(value)
        if row_count == 0:
            raise ParseError("XLSX contains no data rows")
    finally:
        workbook.close()
    columns_info = _build_columns_info(names, accumulators, None, None, sheet_name)
    return ParseResult(
        file_type=ExperimentFileType.XLSX,
        row_count=row_count,
        column_count=column_count,
        columns_info=columns_info,
        encoding=None,
        delimiter=None,
        sheet_name=sheet_name,
    )


def _xls_cell_value(workbook, cell):
    import xlrd

    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
        except (ValueError, OverflowError):
            return str(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return "error"
    return cell.value


def parse_xls(source_path: str | Path) -> ParseResult:
    path = validate_container(source_path, ExperimentFileType.XLS)
    try:
        import xlrd
    except ImportError as exc:
        raise ParseError("XLS parser is unavailable") from exc
    try:
        workbook = xlrd.open_workbook(filename=str(path), on_demand=True, ragged_rows=True)
    except Exception as exc:
        raise ParseError("XLS workbook cannot be parsed", "type") from exc
    try:
        non_empty = [sheet for sheet in workbook.sheets() if sheet.nrows > 0 and sheet.ncols > 0]
        if len(non_empty) != 1:
            raise ParseError("XLS must contain exactly one non-empty worksheet")
        sheet = non_empty[0]
        if sheet.nrows - 1 > _MAX_ROWS:
            raise ParseError("XLS row count exceeds limit", "size")
        if not 1 <= sheet.ncols <= _MAX_COLUMNS:
            raise ParseError("XLS column count exceeds limit", "size")
        names = _validate_column_names([sheet.cell_value(0, index) for index in range(sheet.ncols)])
        accumulators = [_ColumnAccumulator() for _ in names]
        row_count = sheet.nrows - 1
        if row_count == 0:
            raise ParseError("XLS contains no data rows")
        for row_index in range(1, sheet.nrows):
            for column_index, accumulator in enumerate(accumulators):
                cell = sheet.cell(row_index, column_index)
                accumulator.observe(_xls_cell_value(workbook, cell))
        sheet_name = sheet.name
    finally:
        workbook.release_resources()
    columns_info = _build_columns_info(names, accumulators, None, None, sheet_name)
    return ParseResult(
        file_type=ExperimentFileType.XLS,
        row_count=row_count,
        column_count=len(names),
        columns_info=columns_info,
        encoding=None,
        delimiter=None,
        sheet_name=sheet_name,
    )


def parse_experiment_file(
    source_path: str | Path,
    file_type: ExperimentFileType,
) -> ParseResult:
    if file_type == ExperimentFileType.CSV:
        return parse_csv(source_path)
    if file_type == ExperimentFileType.XLSX:
        return parse_xlsx(source_path)
    if file_type == ExperimentFileType.XLS:
        return parse_xls(source_path)
    raise ParseError("unsupported file type", "type")
