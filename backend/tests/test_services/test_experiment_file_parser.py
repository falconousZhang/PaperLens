from __future__ import annotations

import csv
import datetime as dt
import io
import json
import zipfile
from pathlib import Path

import pytest

from paperlens.core.enums import ExperimentFileType
from paperlens.services.experiment_file_parser import (
    ParseError,
    _check_xlsx_zip_safety,
    parse_csv,
    parse_experiment_file,
    parse_xls,
    parse_xlsx,
    validate_container,
    validate_filename_and_type,
)


def _csv_bytes(
    rows: list[list[str]],
    delimiter: str = ",",
    encoding: str = "utf-8",
    bom: bool = False,
) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=delimiter)
    writer.writerows(rows)
    raw = buffer.getvalue().encode(encoding)
    return b"\xef\xbb\xbf" + raw if bom else raw


def _xlsx_bytes(rows: list[list], second_sheet_rows: list[list] | None = None) -> bytes:
    import openpyxl

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    if second_sheet_rows is not None:
        second = workbook.create_sheet("Second")
        for row in second_sheet_rows:
            second.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _xls_bytes(rows: list[list]) -> bytes:
    import xlwt

    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Sheet1")
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            worksheet.write(row_index, column_index, value)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write(tmp_path: Path, filename: str, content: bytes) -> Path:
    path = tmp_path / filename
    path.write_bytes(content)
    return path


def _minimal_zip(entries: dict[str, bytes], compression=zipfile.ZIP_STORED) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=compression) as archive:
        archive.writestr(
            "[Content_Types].xml",
            entries.pop(
                "[Content_Types].xml",
                b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>',
            ),
        )
        archive.writestr("xl/workbook.xml", entries.pop("xl/workbook.xml", b"<workbook/>"))
        for name, content in entries.items():
            archive.writestr(name, content)
    return buffer.getvalue()


@pytest.mark.parametrize(
    ("delimiter", "expected"),
    [(",", ","), (";", ";"), ("\t", "\t")],
)
def test_csv_delimiters_are_deterministic(tmp_path: Path, delimiter: str, expected: str):
    path = _write(tmp_path, "data.csv", _csv_bytes([["name", "score"], ["alice", "9"]], delimiter))
    result = parse_csv(path)
    assert result.delimiter == expected
    assert result.row_count == 1


@pytest.mark.parametrize(
    ("encoding", "bom", "expected"),
    [("utf-8", False, "utf-8"), ("utf-8", True, "utf-8-sig"), ("gb18030", False, "gb18030")],
)
def test_csv_encodings(tmp_path: Path, encoding: str, bom: bool, expected: str):
    path = _write(tmp_path, "中文.csv", _csv_bytes([["名称", "分数"], ["模型", "90"]], encoding=encoding, bom=bom))
    result = parse_csv(path)
    assert result.encoding == expected
    assert result.columns_info["columns"][0]["name"] == "名称"


def test_csv_dtype_null_count_order_and_no_samples(tmp_path: Path):
    raw = _csv_bytes(
        [
            ["int", "float", "bool", "when", "mixed", "empty"],
            ["1", "1.5", "true", "2026-07-14", "1", ""],
            ["2", "2", "false", "2026-07-15T10:11:12", "text", ""],
        ]
    )
    result = parse_csv(_write(tmp_path, "types.csv", raw))
    columns = result.columns_info["columns"]
    assert [column["name"] for column in columns] == ["int", "float", "bool", "when", "mixed", "empty"]
    assert [column["dtype"] for column in columns] == ["integer", "float", "boolean", "datetime", "string", "empty"]
    assert columns[-1]["nullable"] is True
    assert columns[-1]["null_count"] == 2
    assert "2026-07-14" not in json.dumps(result.columns_info)


@pytest.mark.parametrize(
    "raw",
    [b"", b"a,b\n", b"a,b\n1\n", b"a\x00,b\n1,2\n", b'a,b\n"unterminated,2\n'],
)
def test_csv_rejects_empty_or_malformed_content(tmp_path: Path, raw: bytes):
    path = _write(tmp_path, "bad.csv", raw)
    with pytest.raises(ParseError):
        parse_csv(path)


def test_csv_rejects_invalid_encoding(tmp_path: Path):
    path = _write(tmp_path, "bad.csv", b"name\n\xff\xff\xff\x81\n")
    with pytest.raises(ParseError):
        parse_csv(path)


def test_csv_rejects_ambiguous_delimiter(tmp_path: Path):
    path = _write(tmp_path, "bad.csv", b"a,b;c\n1,2;3\n")
    with pytest.raises(ParseError, match="ambiguous"):
        parse_csv(path)


@pytest.mark.parametrize(
    "header",
    [["", "b"], ["a", "a"], ["x" * 129, "b"], ["a\x01", "b"]],
)
def test_csv_rejects_invalid_column_names(tmp_path: Path, header: list[str]):
    path = _write(tmp_path, "bad.csv", _csv_bytes([header, ["1", "2"]]))
    with pytest.raises(ParseError):
        parse_csv(path)


def test_csv_rejects_row_and_column_limits(tmp_path: Path):
    too_many_rows = _write(tmp_path, "rows.csv", b"a\n" + b"1\n" * 100001)
    with pytest.raises(ParseError) as row_error:
        parse_csv(too_many_rows)
    assert row_error.value.kind == "size"
    header = [f"c{index}" for index in range(257)]
    too_many_columns = _write(tmp_path, "columns.csv", _csv_bytes([header, ["1"] * 257]))
    with pytest.raises(ParseError) as column_error:
        parse_csv(too_many_columns)
    assert column_error.value.kind == "size"


def test_xlsx_basic_and_datetime(tmp_path: Path):
    path = _write(
        tmp_path,
        "data.xlsx",
        _xlsx_bytes([["name", "score", "when"], ["alice", 1], ["bob", 2.5, dt.datetime(2026, 7, 14)]]),
    )
    result = parse_xlsx(path)
    assert result.file_type == ExperimentFileType.XLSX
    assert result.row_count == 2
    assert [item["dtype"] for item in result.columns_info["columns"]] == ["string", "float", "datetime"]
    assert result.columns_info["columns"][2]["null_count"] == 1


def test_xlsx_rejects_multiple_non_empty_sheets(tmp_path: Path):
    path = _write(tmp_path, "multi.xlsx", _xlsx_bytes([["a"], [1]], [["b"], [2]]))
    with pytest.raises(ParseError, match="exactly one"):
        parse_xlsx(path)


def test_xlsx_allows_one_non_empty_and_one_empty_sheet(tmp_path: Path):
    path = _write(tmp_path, "one.xlsx", _xlsx_bytes([["a"], [1]], []))
    assert parse_xlsx(path).row_count == 1


@pytest.mark.parametrize(
    "entry_name",
    ["../escape", "xl/../escape", "..\\escape", "C:/escape", "/absolute"],
)
def test_xlsx_rejects_entry_path_traversal(tmp_path: Path, entry_name: str):
    path = _write(tmp_path, "unsafe.xlsx", _minimal_zip({entry_name: b"bad"}))
    with pytest.raises(ParseError, match="unsafe"):
        parse_xlsx(path)


def test_xlsx_rejects_duplicate_entries(tmp_path: Path):
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
        archive.writestr("xl/workbook.xml", "<workbook/>")
        archive.writestr("xl/workbook.xml", "<workbook/>")
    path = _write(tmp_path, "duplicate.xlsx", buffer.getvalue())
    with pytest.raises(ParseError, match="duplicate"):
        parse_xlsx(path)


def test_xlsx_rejects_per_entry_zip_bomb(tmp_path: Path):
    content = b"0" * (2 * 1024 * 1024)
    path = _write(
        tmp_path,
        "bomb.xlsx",
        _minimal_zip({"xl/media/bomb.bin": content}, compression=zipfile.ZIP_DEFLATED),
    )
    with pytest.raises(ParseError) as error:
        parse_xlsx(path)
    assert error.value.kind == "size"


def test_xlsx_rejects_encrypted_entry_flag():
    buffer = io.BytesIO(_minimal_zip({}))
    archive = zipfile.ZipFile(buffer)
    archive.infolist()[0].flag_bits |= 0x1
    with pytest.raises(ParseError, match="encrypted"):
        _check_xlsx_zip_safety(archive)
    archive.close()


@pytest.mark.parametrize(
    "entry_name",
    ["xl/vbaProject.bin", "xl/externalLinks/externalLink1.xml", "xl/embeddings/oleObject1.bin", "xl/oleObjects/oleObject1.bin", "xl/activeX/activeX1.bin"],
)
def test_xlsx_rejects_active_content_entries(tmp_path: Path, entry_name: str):
    path = _write(tmp_path, "active.xlsx", _minimal_zip({entry_name: b"active"}))
    with pytest.raises(ParseError):
        parse_xlsx(path)


def test_xlsx_rejects_macro_content_type(tmp_path: Path):
    content_types = b'<Types><Override ContentType="application/vnd.ms-excel.sheet.macroEnabled.main+xml"/></Types>'
    path = _write(tmp_path, "macro.xlsx", _minimal_zip({"[Content_Types].xml": content_types}))
    with pytest.raises(ParseError, match="macros"):
        parse_xlsx(path)


def test_xlsx_rejects_external_relationship(tmp_path: Path):
    relationship = b'<Relationships><Relationship TargetMode="External" Target="https://example.invalid"/></Relationships>'
    path = _write(tmp_path, "external.xlsx", _minimal_zip({"xl/_rels/workbook.xml.rels": relationship}))
    with pytest.raises(ParseError, match="external"):
        parse_xlsx(path)


@pytest.mark.parametrize("formula_xml", [b"<worksheet><f>SUM(A1:A2)</f></worksheet>", b"<worksheet><f/></worksheet>", b'<x:worksheet xmlns:x="urn:x"><x:f/></x:worksheet>'])
def test_xlsx_rejects_all_formula_element_forms(tmp_path: Path, formula_xml: bytes):
    path = _write(tmp_path, "formula.xlsx", _minimal_zip({"xl/worksheets/sheet1.xml": formula_xml}))
    with pytest.raises(ParseError, match="formula"):
        parse_xlsx(path)


@pytest.mark.parametrize(
    "content",
    [b"not zip", _minimal_zip({"xl/workbook.xml": b"<workbook/>"}).replace(b"[Content_Types].xml", b"Missing_Content_Type")],
)
def test_xlsx_rejects_invalid_container(tmp_path: Path, content: bytes):
    path = _write(tmp_path, "invalid.xlsx", content)
    with pytest.raises(ParseError):
        parse_xlsx(path)


def test_xls_basic_and_ole_validation(tmp_path: Path):
    path = _write(tmp_path, "data.xls", _xls_bytes([["name", "score"], ["alice", 1], ["bob", 2.5]]))
    result = parse_xls(path)
    assert result.file_type == ExperimentFileType.XLS
    assert result.row_count == 2
    assert [item["dtype"] for item in result.columns_info["columns"]] == ["string", "float"]
    invalid = _write(tmp_path, "invalid.xls", b"not ole")
    with pytest.raises(ParseError) as error:
        parse_xls(invalid)
    assert error.value.kind == "type"


def test_xls_rejects_header_only(tmp_path: Path):
    path = _write(tmp_path, "empty.xls", _xls_bytes([["a", "b"]]))
    with pytest.raises(ParseError, match="no data"):
        parse_xls(path)


@pytest.mark.parametrize(
    ("filename", "expected"),
    [("data.csv", ExperimentFileType.CSV), ("DATA.XLSX", ExperimentFileType.XLSX), (r"C:\\fakepath\\data.xls", ExperimentFileType.XLS)],
)
def test_filename_type_and_basename(filename: str, expected: ExperimentFileType):
    basename, file_type = validate_filename_and_type(filename)
    assert "/" not in basename and "\\" not in basename
    assert file_type == expected


@pytest.mark.parametrize("filename", ["", "data.txt", "data.xlsm", "data.xlsb", "a\x01.csv", "a\n.csv", "a\x7f.csv", f"{'a' * 252}.csv"])
def test_filename_rejections(filename: str):
    with pytest.raises(ParseError) as error:
        validate_filename_and_type(filename)
    assert error.value.kind == "type"


@pytest.mark.parametrize(
    ("filename", "content", "file_type"),
    [
        ("fake.xlsx", b"a,b\n1,2\n", ExperimentFileType.XLSX),
        ("fake.xls", b"a,b\n1,2\n", ExperimentFileType.XLS),
        ("fake.csv", _xlsx_bytes([["a"], [1]]), ExperimentFileType.CSV),
        ("fake.csv", b"\x1f\x8bcompressed", ExperimentFileType.CSV),
    ],
)
def test_extension_magic_mismatch(tmp_path: Path, filename: str, content: bytes, file_type: ExperimentFileType):
    path = _write(tmp_path, filename, content)
    with pytest.raises(ParseError) as error:
        validate_container(path, file_type)
    assert error.value.kind == "type"


@pytest.mark.parametrize(
    ("filename", "content", "file_type"),
    [
        ("data.csv", _csv_bytes([["a"], ["1"]]), ExperimentFileType.CSV),
        ("data.xlsx", _xlsx_bytes([["a"], [1]]), ExperimentFileType.XLSX),
        ("data.xls", _xls_bytes([["a"], [1]]), ExperimentFileType.XLS),
    ],
)
def test_parser_dispatch_uses_server_path(tmp_path: Path, filename: str, content: bytes, file_type: ExperimentFileType):
    path = _write(tmp_path, filename, content)
    result = parse_experiment_file(path, file_type)
    assert result.file_type == file_type
